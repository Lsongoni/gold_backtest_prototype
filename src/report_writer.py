from __future__ import annotations

import logging
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import (
    DETAIL_COLUMNS,
    MARKET_SUMMARY_COLUMNS,
    MONTHLY_SUMMARY_COLUMNS,
    STRATEGY_RANKING_COLUMNS,
    WARNING_COLUMNS,
)

logger = logging.getLogger(__name__)


def _ensure_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = pd.NA
    return out[columns]


def _dedupe_warnings(warnings: list[dict]) -> pd.DataFrame:
    if not warnings:
        return pd.DataFrame(columns=WARNING_COLUMNS)
    df = pd.DataFrame(warnings)
    for col in WARNING_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[WARNING_COLUMNS].drop_duplicates().reset_index(drop=True)


def _style_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells[:2000]:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            width = min(max(max_length + 2, 10), 42)
            worksheet.column_dimensions[get_column_letter(column)].width = width


def _placeholder_chart(path: Path, title: str, message: str) -> None:
    plt.figure(figsize=(10, 5))
    plt.title(title)
    plt.text(0.5, 0.5, message, ha="center", va="center", fontsize=12)
    plt.xticks([])
    plt.yticks([])
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()


def generate_charts(
    detail: pd.DataFrame,
    monthly_summary: pd.DataFrame,
    ranking: pd.DataFrame,
    chart_dir: str | Path,
) -> None:
    chart_path = Path(chart_dir)
    chart_path.mkdir(parents=True, exist_ok=True)

    cumulative_file = chart_path / "cumulative_pnl_best_strategy.png"
    monthly_file = chart_path / "monthly_pnl_best_strategy.png"
    distribution_file = chart_path / "spread_distribution_best_strategy.png"

    valid_ranking = ranking.dropna(subset=["score"]) if not ranking.empty else pd.DataFrame()
    if valid_ranking.empty:
        _placeholder_chart(cumulative_file, "Cumulative PnL", "No valid strategy result")
        _placeholder_chart(monthly_file, "Monthly PnL", "No valid strategy result")
        _placeholder_chart(distribution_file, "Spread Distribution", "No valid strategy result")
        return

    best_strategy = valid_ranking.iloc[0]["strategy_id"]
    best_detail = detail.loc[(detail["strategy_id"] == best_strategy) & detail["pnl_amount"].notna()].copy()
    if best_detail.empty:
        _placeholder_chart(cumulative_file, "Cumulative PnL", "No valid PnL rows")
    else:
        best_detail = best_detail.sort_values(["date", "order_datetime"])
        best_detail["cumulative_pnl"] = best_detail["pnl_amount"].fillna(0).cumsum()
        plt.figure(figsize=(11, 5))
        plt.plot(best_detail["order_datetime"], best_detail["cumulative_pnl"], linewidth=1.8)
        plt.title(f"Cumulative PnL - {best_strategy}")
        plt.xlabel("Date")
        plt.ylabel("Cumulative PnL")
        plt.grid(True, alpha=0.25)
        plt.tight_layout()
        plt.savefig(cumulative_file, dpi=160)
        plt.close()

    best_monthly = monthly_summary.loc[monthly_summary["strategy_id"] == best_strategy].copy()
    best_monthly = best_monthly.dropna(subset=["monthly_pnl"])
    if best_monthly.empty:
        _placeholder_chart(monthly_file, "Monthly PnL", "No valid monthly PnL rows")
    else:
        best_monthly = best_monthly.sort_values("month")
        month_values = pd.to_datetime(best_monthly["month"], errors="coerce")
        plt.figure(figsize=(11, 5))
        plt.bar(month_values, best_monthly["monthly_pnl"], width=20)
        plt.title(f"Monthly PnL - {best_strategy}")
        plt.xlabel("Month")
        plt.ylabel("Monthly PnL")
        plt.xticks(month_values, best_monthly["month"], rotation=45, ha="right")
        plt.grid(True, axis="y", alpha=0.25)
        plt.tight_layout()
        plt.savefig(monthly_file, dpi=160)
        plt.close()

    valid_spread = detail.loc[
        (detail["strategy_id"] == best_strategy) & detail["spread_per_gram"].notna(),
        "spread_per_gram",
    ]
    if valid_spread.empty:
        _placeholder_chart(distribution_file, "Spread Distribution", "No valid spread rows")
    else:
        plt.figure(figsize=(10, 5))
        plt.hist(valid_spread, bins=30, edgecolor="white")
        plt.axvline(6, color="red", linestyle="--", linewidth=1.5, label="Target 6")
        plt.title(f"Spread Distribution - {best_strategy}")
        plt.xlabel("Spread per gram")
        plt.ylabel("Count")
        plt.legend()
        plt.grid(True, axis="y", alpha=0.25)
        plt.tight_layout()
        plt.savefig(distribution_file, dpi=160)
        plt.close()

    logger.info("Charts written to %s", chart_path)


def write_report(
    output_dir: str | Path,
    orders_clean: pd.DataFrame,
    market_data_summary: pd.DataFrame,
    daily_detail: pd.DataFrame,
    monthly_summary: pd.DataFrame,
    strategy_ranking: pd.DataFrame,
    warnings: list[dict],
) -> Path:
    output_path = Path(output_dir)
    chart_dir = output_path / "charts"
    output_path.mkdir(parents=True, exist_ok=True)
    chart_dir.mkdir(parents=True, exist_ok=True)

    daily_detail = _ensure_columns(daily_detail, DETAIL_COLUMNS)
    market_data_summary = _ensure_columns(market_data_summary, MARKET_SUMMARY_COLUMNS)
    monthly_summary = _ensure_columns(monthly_summary, MONTHLY_SUMMARY_COLUMNS)
    strategy_ranking = _ensure_columns(strategy_ranking, STRATEGY_RANKING_COLUMNS)
    warnings_df = _dedupe_warnings(warnings)

    generate_charts(daily_detail, monthly_summary, strategy_ranking, chart_dir)

    excel_path = output_path / "backtest_results.xlsx"
    logger.info("Writing Excel report: %s", excel_path)
    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        datetime_format="yyyy-mm-dd hh:mm:ss",
        date_format="yyyy-mm-dd",
    ) as writer:
        orders_clean.to_excel(writer, sheet_name="orders_clean", index=False)
        market_data_summary.to_excel(writer, sheet_name="market_data_summary", index=False)
        daily_detail.to_excel(writer, sheet_name="daily_detail", index=False)
        monthly_summary.to_excel(writer, sheet_name="monthly_summary", index=False)
        strategy_ranking.to_excel(writer, sheet_name="strategy_ranking", index=False)
        warnings_df.to_excel(writer, sheet_name="warnings", index=False)
        _style_workbook(writer)

    return excel_path
